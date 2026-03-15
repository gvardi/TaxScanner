"""Excel report generation using openpyxl."""

import os
from datetime import datetime
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation

from taxscanner.classifier.models import ClassificationResult
from taxscanner.utils.logging import get_logger

logger = get_logger(__name__)

# Styles
HEADER_FONT = Font(bold=True, color="FFFFFF", size=11)
HEADER_FILL = PatternFill(start_color="2F5496", end_color="2F5496", fill_type="solid")
RED_FILL = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
YELLOW_FILL = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")
GREEN_FILL = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
THIN_BORDER = Border(
    left=Side(style="thin"),
    right=Side(style="thin"),
    top=Side(style="thin"),
    bottom=Side(style="thin"),
)

LOW_CONFIDENCE_THRESHOLD = 0.5
REVIEW_CONFIDENCE_THRESHOLD = 0.75

EXPENSE_HEADERS = [
    "Date", "Vendor", "Amount", "Category", "Business",
    "AI Confidence", "AI Reasoning", "Status", "Reviewed By",
    "Notes", "Source", "Gmail Link",
]


def _style_header_row(ws, headers: list[str]):
    """Apply styling to the header row."""
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = Alignment(horizontal="center", wrap_text=True)
        cell.border = THIN_BORDER

    # Freeze the header row
    ws.freeze_panes = "A2"

    # Auto-filter
    ws.auto_filter.ref = f"A1:{get_column_letter(len(headers))}1"


def _auto_width(ws, min_width: int = 10, max_width: int = 50):
    """Auto-adjust column widths based on content."""
    for col in ws.columns:
        max_len = 0
        col_letter = get_column_letter(col[0].column)
        for cell in col:
            if cell.value:
                max_len = max(max_len, len(str(cell.value)))
        adjusted = min(max(max_len + 2, min_width), max_width)
        ws.column_dimensions[col_letter].width = adjusted


def _write_expense_row(ws, row: int, c: ClassificationResult):
    """Write a classification result as an expense row."""
    ws.cell(row=row, column=1, value=c.invoice_date if c.invoice_date != "Unknown" else c.date)
    ws.cell(row=row, column=2, value=c.vendor)
    ws.cell(row=row, column=3, value=c.amount)
    ws.cell(row=row, column=4, value=c.category)
    ws.cell(row=row, column=5, value=c.business)

    conf_cell = ws.cell(row=row, column=6, value=round(c.confidence, 2))
    if c.confidence < LOW_CONFIDENCE_THRESHOLD:
        conf_cell.fill = RED_FILL
    elif c.confidence < REVIEW_CONFIDENCE_THRESHOLD:
        conf_cell.fill = YELLOW_FILL
    else:
        conf_cell.fill = GREEN_FILL

    ws.cell(row=row, column=7, value=c.reasoning)
    ws.cell(row=row, column=8, value="Needs Review")
    ws.cell(row=row, column=9, value="")
    ws.cell(row=row, column=10, value="")
    ws.cell(row=row, column=11, value=c.source)
    ws.cell(row=row, column=12, value=c.gmail_link)

    # Style all cells in the row
    for col in range(1, 13):
        ws.cell(row=row, column=col).border = THIN_BORDER


def _add_status_validation(ws, last_row: int):
    """Add dropdown data validation for the Status column."""
    if last_row < 2:
        return
    dv = DataValidation(
        type="list",
        formula1='"Approved,Rejected,Needs Review"',
        allow_blank=True,
    )
    dv.error = "Please select a valid status"
    dv.errorTitle = "Invalid Status"
    ws.add_data_validation(dv)
    dv.add(f"H2:H{last_row}")


def _create_expenses_sheet(wb: Workbook, classifications: list[ClassificationResult], sheet_name: str):
    """Create an expenses sheet (used for both Business and Personal)."""
    ws = wb.create_sheet(title=sheet_name)
    _style_header_row(ws, EXPENSE_HEADERS)

    for i, c in enumerate(classifications, 2):
        _write_expense_row(ws, i, c)

    _add_status_validation(ws, len(classifications) + 1)
    _auto_width(ws)

    return ws


def _parse_amount(amount_str: str) -> float | None:
    """Parse a currency string like '$1,234.56' into a float, or None on failure."""
    try:
        return float(amount_str.replace("$", "").replace(",", "").strip())
    except (ValueError, AttributeError):
        return None


def _write_summary_group(ws, start_row: int, title: str, header_label: str,
                         groups: dict[str, list[ClassificationResult]]) -> int:
    """Write a grouped summary section and return the next available row."""
    ws.cell(row=start_row, column=1, value=title).font = Font(bold=True, size=12)
    header_row = start_row + 1
    ws.cell(row=header_row, column=1, value=header_label)
    ws.cell(row=header_row, column=2, value="Count")
    ws.cell(row=header_row, column=3, value="Expenses")

    for col in range(1, 4):
        ws.cell(row=header_row, column=col).font = HEADER_FONT
        ws.cell(row=header_row, column=col).fill = HEADER_FILL
        ws.cell(row=header_row, column=col).border = THIN_BORDER

    row = header_row + 1
    for label, items in sorted(groups.items()):
        ws.cell(row=row, column=1, value=label).border = THIN_BORDER
        ws.cell(row=row, column=2, value=len(items)).border = THIN_BORDER

        amounts = [a for item in items if (a := _parse_amount(item.amount)) is not None]
        if amounts:
            ws.cell(row=row, column=3, value=f"${sum(amounts):,.2f}").border = THIN_BORDER
        else:
            ws.cell(row=row, column=3, value="N/A").border = THIN_BORDER
        row += 1

    return row


def _create_summary_sheet(wb: Workbook, business_expenses: list[ClassificationResult]):
    """Create the Summary sheet with totals by business and category."""
    ws = wb.create_sheet(title="Summary")

    ws.cell(row=1, column=1, value="Expense Summary").font = Font(bold=True, size=14)

    # By business
    business_groups: dict[str, list] = {}
    for c in business_expenses:
        business_groups.setdefault(c.business or "Unassigned", []).append(c)
    row = _write_summary_group(ws, 3, "By Business", "Business", business_groups)

    # By category
    category_groups: dict[str, list] = {}
    for c in business_expenses:
        category_groups.setdefault(c.category, []).append(c)
    row = _write_summary_group(ws, row + 2, "By Category", "Category", category_groups)

    # Needs review count
    row += 2
    needs_review = [c for c in business_expenses if c.confidence < REVIEW_CONFIDENCE_THRESHOLD]
    ws.cell(row=row, column=1, value="Items needing review:").font = Font(bold=True)
    ws.cell(row=row, column=2, value=len(needs_review))

    _auto_width(ws)


def _create_skipped_sheet(wb: Workbook, skipped: list[dict]):
    """Create the Skipped sheet."""
    ws = wb.create_sheet(title="Skipped")

    headers = ["Email ID", "Subject", "From", "Date", "Reason"]
    _style_header_row(ws, headers)

    for i, item in enumerate(skipped, 2):
        ws.cell(row=i, column=1, value=item.get("id", "")).border = THIN_BORDER
        ws.cell(row=i, column=2, value=item.get("subject", "")).border = THIN_BORDER
        ws.cell(row=i, column=3, value=item.get("from", "")).border = THIN_BORDER
        ws.cell(row=i, column=4, value=item.get("date", "")).border = THIN_BORDER
        ws.cell(row=i, column=5, value=item.get("reason", "")).border = THIN_BORDER

    _auto_width(ws)


def generate_report(
    classifications: list[ClassificationResult],
    skipped: list[dict],
    year: int,
    config,
) -> str:
    """Generate the Excel report and return the file path."""
    output_dir = (
        config.report.output_dir
        if hasattr(config, "report") and hasattr(config.report, "output_dir")
        else config.get("report", {}).get("output_dir", "reports")
    )
    Path(output_dir).mkdir(parents=True, exist_ok=True)

    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    filename = f"tax_scan_{year}_{timestamp}.xlsx"
    filepath = os.path.join(output_dir, filename)

    wb = Workbook()
    # Remove default sheet
    wb.remove(wb.active)

    # Split classifications
    business = [c for c in classifications if c.expense_type == "business"]
    personal = [c for c in classifications if c.expense_type == "personal"]
    uncertain = [c for c in classifications if c.expense_type == "uncertain"]

    # Business expenses go in main sheet, uncertain ones too (for review)
    main_expenses = business + uncertain

    # Sheet 1: Expenses (business + uncertain)
    _create_expenses_sheet(wb, main_expenses, "Expenses")

    # Sheet 2: Summary
    _create_summary_sheet(wb, main_expenses)

    # Sheet 3: Personal
    _create_personal_sheet(wb, personal)

    # Sheet 4: Skipped
    _create_skipped_sheet(wb, skipped)

    wb.save(filepath)
    logger.info(f"Report saved: {filepath}")

    return filepath


def _create_personal_sheet(wb: Workbook, personal: list[ClassificationResult]):
    """Create the Personal expenses sheet."""
    _create_expenses_sheet(wb, personal, "Personal")
