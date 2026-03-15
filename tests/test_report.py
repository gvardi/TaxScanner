"""Tests for report generation."""

import os
import pytest
from openpyxl import load_workbook

from taxscanner.classifier.models import ClassificationResult
from taxscanner.report.excel import generate_report


@pytest.fixture
def sample_classifications():
    return [
        ClassificationResult(
            message_id="msg1",
            subject="AWS Invoice",
            sender="aws@amazon.com",
            date="2025-03-01",
            gmail_link="https://mail.google.com/mail/u/0/#inbox/msg1",
            source="pdf:invoice.pdf",
            expense_type="business",
            business="Software/IT",
            category="Cloud & Hosting",
            vendor="Amazon Web Services",
            amount="$150.00",
            invoice_date="2025-03-01",
            confidence=0.95,
            reasoning="AWS hosting",
        ),
        ClassificationResult(
            message_id="msg2",
            subject="Home Depot Receipt",
            sender="homedepot@email.com",
            date="2025-04-15",
            gmail_link="https://mail.google.com/mail/u/0/#inbox/msg2",
            source="body",
            expense_type="business",
            business="Real Estate",
            category="Property Maintenance & Repairs",
            vendor="Home Depot",
            amount="$89.50",
            invoice_date="2025-04-15",
            confidence=0.88,
            reasoning="Home improvement supplies for rental property",
        ),
        ClassificationResult(
            message_id="msg3",
            subject="Netflix Subscription",
            sender="info@netflix.com",
            date="2025-02-01",
            gmail_link="https://mail.google.com/mail/u/0/#inbox/msg3",
            source="body",
            expense_type="personal",
            business="",
            category="Software & Subscriptions",
            vendor="Netflix",
            amount="$15.99",
            invoice_date="2025-02-01",
            confidence=0.98,
            reasoning="Personal entertainment subscription",
        ),
        ClassificationResult(
            message_id="msg4",
            subject="Amazon Order",
            sender="orders@amazon.com",
            date="2025-05-10",
            gmail_link="https://mail.google.com/mail/u/0/#inbox/msg4",
            source="body",
            expense_type="uncertain",
            business="",
            category="Other",
            vendor="Amazon",
            amount="$45.00",
            invoice_date="2025-05-10",
            confidence=0.35,
            reasoning="Could be personal or business, unclear from description",
        ),
    ]


@pytest.fixture
def sample_skipped():
    return [
        {
            "id": "skip1",
            "subject": "Newsletter",
            "from": "news@example.com",
            "date": "2025-01-01",
            "reason": "No extractable invoice data",
        },
    ]


@pytest.fixture
def config():
    return {"report": {"output_dir": "test_reports"}}


def test_generate_report(sample_classifications, sample_skipped, config, tmp_path):
    config["report"]["output_dir"] = str(tmp_path)
    filepath = generate_report(sample_classifications, sample_skipped, 2025, config)

    assert os.path.exists(filepath)
    assert "tax_scan_2025" in filepath

    wb = load_workbook(filepath)
    sheet_names = wb.sheetnames

    assert "Expenses" in sheet_names
    assert "Summary" in sheet_names
    assert "Personal" in sheet_names
    assert "Skipped" in sheet_names

    # Expenses sheet: business (2) + uncertain (1) = 3 rows + header
    expenses_ws = wb["Expenses"]
    assert expenses_ws.max_row == 4  # 1 header + 3 data rows

    # Personal sheet: 1 personal + header
    personal_ws = wb["Personal"]
    assert personal_ws.max_row == 2  # 1 header + 1 data row

    # Skipped sheet: 1 skipped + header
    skipped_ws = wb["Skipped"]
    assert skipped_ws.max_row == 2

    # Verify header
    assert expenses_ws.cell(row=1, column=1).value == "Date"
    assert expenses_ws.cell(row=1, column=2).value == "Vendor"

    # Verify data
    assert expenses_ws.cell(row=2, column=2).value == "Amazon Web Services"
    assert expenses_ws.cell(row=2, column=3).value == "$150.00"

    wb.close()


def test_empty_report(config, tmp_path):
    config["report"]["output_dir"] = str(tmp_path)
    filepath = generate_report([], [], 2025, config)

    assert os.path.exists(filepath)
    wb = load_workbook(filepath)
    assert "Expenses" in wb.sheetnames
    wb.close()
